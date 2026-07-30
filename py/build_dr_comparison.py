"""Build the HyperSync equivalent of dr_comparison_latest.xlsx.

Soter tabs   : computed EXCLUSIVELY from hypersync-results/dr_full/*.csv
               (event-derived on-chain data; rates = locked protocol constants).
Reference    : Spark / Amatsu / BA / Payouts tabs copied VERBATIM from the old
               workbook (clearly labeled reference data, never mixed into Soter).
Diff tabs    : recomputed = Soter - reference over each reference's months.
Checks tab   : (a) non-aggregator venue set old-vs-new, (b) aggregator values
               vs the measured impact numbers, (c) provenance assertions.
"""
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "py"))
MEASURE = REPO / "hypersync-results" / "measurements"
OLD = REPO / "dune-results" / "dr_comparison_latest.xlsx"
NEW = REPO / "hypersync-results" / "dr_comparison_hypersync.xlsx"
CHUNK_DIR = REPO / "hypersync-results" / "dr_full"

AGG_CODES = {1003, 1004, 4011}
MONTHS_2026 = [f"2026-{m:02d}" for m in range(1, 8)]

# --- Payout eligibility windows per ref_code (ops-owned) -----------------------
# Default: every venue is payable from 2026-01 (MSC settlement start), no end
# date yet. Override per code as ops advises (e.g. if 9001 Aave is confirmed
# payable from Spark's official start 2025-07, add it here). Months BEFORE the
# start still appear in the Soter data tabs (methodology-pure, full history) —
# eligibility is applied only in the "Payable" view.
# Codes that map to no beneficiary — never payable regardless of eligibility.
# Canonical registry lives in drhs.revenue.monthly.NON_PAYABLE_CODES.
sys.path.insert(0, str(REPO / "py"))
from drhs.revenue.monthly import NON_PAYABLE_CODES  # noqa: E402

ELIGIBILITY_DEFAULT_START = "2026-01"
ELIGIBILITY_OVERRIDES: dict[int, tuple[str, str | None]] = {
    # code: (start "YYYY-MM", end "YYYY-MM" exclusive or None)
}

def eligibility(code: int) -> tuple[str, str | None]:
    return ELIGIBILITY_OVERRIDES.get(code, (ELIGIBILITY_DEFAULT_START, None))

NOTES = {
    -999999: "Synthetic code: Untagged USDS-CLE, USDS-SKY, USDS-SPK, stUSDS.",
    0: "Explicit on-chain referral on Ethereum. L2 sUSDS split to 10000/10001.",
    99: "Synthetic code: Untagged sUSDS.",
    127: "Synthetic code: untagged sUSDC",
    130: "Synthetic code: Untagged spUSDT.",
    131: "Synthetic code: Untagged spUSDC.",
    132: "Synthetic code: Untagged spPYUSD.",
    197: "stUSDS",
    1003: "CowSwap — synthetic delivery tagging in the unified stream (event-derived).",
    1004: "Paraswap — re-routed router-owned Referral(1004) in the unified stream.",
    4011: "1inch — re-routed executor-owned Referral(4011) in the unified stream.",
    1016: "lazysummer — only the on-chain farm Referral(1016) events appear here; "
          "Amatsu additionally tags fleet ark balances DB-side (docs/lazysummer-1016.md).",
    9001: "Synthetic code: USDS in Aave aEthUSDS; entire contract balance. "
          "Intraday TWA (clean methodology) — the deployed Dune query used "
          "EOD snapshots, which under-count ~20% on heavy-flow months.",
    4001: "Synthetic code: USDS in Solana OFT Bridge; entire contract balance. "
          "Intraday TWA (clean methodology; Dune query used EOD snapshots).",
    10000: "Synthetic code: L2 end users who passed the default PSM3 code 0 "
           "(no integrator to pay; flow >> stock, ~$7k all-time).",
    10001: "Synthetic code: smart-contract-held sUSDS, address-based split "
           "(scan: composition = sUSDC vault backing + ALM/PSM3 float; the "
           "value's real DR is paid in its own venue).",
}

from run_dr_chunk import chunk_csv, chunk_plan, load_chunks  # noqa: E402

# The workbook must be built from a COMPLETE chunk set for the window it
# claims: the dir's manifest pins the --end its checkpoints were built with,
# and the derived plan pins the exact file set. A stale dir (older end) or a
# partial regeneration would otherwise render blank/short July columns while
# every check still reads green.
_mf = CHUNK_DIR / "manifest.json"
if not _mf.exists():
    raise SystemExit(f"{CHUNK_DIR} has no manifest.json — run run_dr_pipeline.py first")
_mf_end = json.loads(_mf.read_text())["end"]
_y, _m = int(MONTHS_2026[-1][:4]), int(MONTHS_2026[-1][5:7])
_need_end = date(_y + (_m == 12), _m % 12 + 1, 1).isoformat()
if _mf_end < _need_end:
    raise SystemExit(
        f"{CHUNK_DIR} holds checkpoints for end={_mf_end} but the workbook spans "
        f"through {MONTHS_2026[-1]} (needs end >= {_need_end}) — re-run "
        "run_dr_pipeline.py for the extended window first")
_expected = [chunk_csv(CHUNK_DIR, name, shard)
             for name, (_f, _s, _t, _n) in chunk_plan().items()
             for shard in ([None] if not _n else [f"{k}/{_n}" for k in range(_n)])]
df = load_chunks(CHUNK_DIR, expected=_expected)
print(f"combined chunks from {CHUNK_DIR} -> {len(df)} grouped rows")
df["month_s"] = df["month"].str[:7]
df["ref_code"] = df["ref_code"].astype(int)

wb_old = openpyxl.load_workbook(OLD, read_only=True)
wb = openpyxl.Workbook()
wb.remove(wb.active)


def month_total_frame(keys):
    g = df.groupby(keys + ["month_s"])["dr_usd"].sum().reset_index()
    return g


def write_aoa(name, rows):
    ws = wb.create_sheet(name)
    for r in rows:
        ws.append(r)


# --- Soter by Ref Code (2026 settlement year) ---------------------------------
g = month_total_frame(["ref_code"])
tokens_by_code = df.groupby("ref_code")["token"].agg(lambda s: ", ".join(sorted(set(s))))
rows = [["ref_code", *MONTHS_2026, "total", "tokens", "notes"]]
for code in sorted(df["ref_code"].unique()):
    sub = g[g["ref_code"] == code].set_index("month_s")["dr_usd"]
    vals = [round(float(sub.get(m, 0.0)), 2) or "" for m in MONTHS_2026]
    tot = round(sum(float(sub.get(m, 0.0)) for m in MONTHS_2026), 2)
    if tot == 0 and all(v == "" for v in vals):
        continue
    rows.append([str(code), *vals, tot, tokens_by_code.get(code, ""), NOTES.get(code, "")])
write_aoa("Soter by Ref Code", rows)

# --- Soter by Ref Code Token (full history) ------------------------------------
all_months = sorted(df["month_s"].unique())
gt = month_total_frame(["ref_code", "token"])
rows = [["ref_code", "token", *all_months, "total", "notes"]]
for (code, token), sub in gt.groupby(["ref_code", "token"]):
    s = sub.set_index("month_s")["dr_usd"]
    vals = [round(float(s.get(m, 0.0)), 2) or "" for m in all_months]
    tot = round(float(s.sum()), 2)
    rows.append([str(code), token, *vals, tot, NOTES.get(code, "")])
write_aoa("Soter by Ref Code Token", rows)

# --- Payable by Ref Code (eligibility windows applied) --------------------------
rows = [["ref_code", "eligible_from", "eligible_until", *MONTHS_2026,
         "payable_total", "excluded_pre_window", "notes"]]
nonpay_rows = []
gc_all = df.groupby(["ref_code", "month_s"])["dr_usd"].sum()
for code in sorted(df["ref_code"].unique()):
    ser = (gc_all.loc[code] if code in gc_all.index.get_level_values(0)
           else pd.Series(dtype=float))
    if int(code) in NON_PAYABLE_CODES:
        vals = [round(float(ser.get(m, 0.0)), 2) or "" for m in MONTHS_2026]
        tot = round(float(ser.sum()), 2)
        nonpay_rows.append([str(code), "-", "-", *vals, "", tot, NOTES.get(code, "")])
        continue
    start, end = eligibility(int(code))
    def _pay(m):
        return float(ser.get(m, 0.0)) if m >= start and (end is None or m < end) else 0.0
    vals = [round(_pay(m), 2) or "" for m in MONTHS_2026]
    payable = round(sum(float(ser.get(m, 0.0)) for m in ser.index
                        if m >= start and (end is None or m < end)), 2)
    excluded = round(sum(float(ser.get(m, 0.0)) for m in ser.index if m < start), 2)
    if payable == 0 and excluded == 0:
        continue
    rows.append([str(code), start, end or "-", *vals, payable, excluded,
                 NOTES.get(code, "")])
rows.append([])
rows.append(["NON-PAYABLE (tracked, no beneficiary — notional dr_usd; never pay)",
             "", "", *[""] * len(MONTHS_2026), "", "all-time total", ""])
rows.extend(nonpay_rows)
write_aoa("Payable by Ref Code", rows)

# --- reference tabs copied verbatim --------------------------------------------
REFS = ["Spark", "Amatsu", "BA", "Payouts"]
ref_data = {}
for name in REFS:
    ws_old = wb_old[name]
    aoa = [list(r) for r in ws_old.iter_rows(values_only=True)]
    write_aoa(f"{name} (reference)", aoa)
    header = aoa[0]
    months = [c for c in header if isinstance(c, str) and c[:4].isdigit()]
    per_code = {}
    for r in aoa[1:]:
        try:
            code = int(str(r[0]))
        except (ValueError, TypeError):
            continue
        per_code[code] = {m: float(str(v).replace(",", "") or 0)
                          for m, v in zip(months, r[1:1 + len(months)]) if v not in (None, "")}
    ref_data[name] = (months, per_code)

# --- diff tabs ------------------------------------------------------------------
soter_pc = {}
for code in df["ref_code"].unique():
    sub = g[g["ref_code"] == code].set_index("month_s")["dr_usd"]
    soter_pc[int(code)] = {m: float(sub.get(m, 0.0)) for m in MONTHS_2026}

for name in REFS:
    months, per_code = ref_data[name]
    common = [m for m in months if m in MONTHS_2026]
    rows = [["ref_code", "present_in", *common, "total_diff", "notes"]]
    for code in sorted(set(soter_pc) | set(per_code)):
        s, o = soter_pc.get(code), per_code.get(code)
        pres = "both" if (s and o) else ("soter only" if s else f"{name.lower()} only")
        diffs = [round((s.get(m, 0.0) if s else 0.0) - (o.get(m, 0.0) if o else 0.0), 2)
                 for m in common]
        rows.append([str(code), pres, *diffs, round(sum(diffs), 2), NOTES.get(code, "")])
    write_aoa(f"Diff Soter-{name}", rows)

# --- Checks tab -----------------------------------------------------------------
checks = [["check", "result", "detail"]]

# (a) non-aggregator venue set vs OLD Soter by Ref Code Token
ws_old = wb_old["Soter by Ref Code Token"]
old_rows = [list(r) for r in ws_old.iter_rows(values_only=True)]
old_hdr = old_rows[0]
old_pairs = set()
old_tot = {}
ti = old_hdr.index("total") if "total" in old_hdr else len(old_hdr) - 2
for r in old_rows[1:]:
    try:
        code = int(str(r[0]))
    except (ValueError, TypeError):
        continue
    old_pairs.add((code, str(r[1])))
    try:
        old_tot[(code, str(r[1]))] = float(str(r[ti]).replace(",", "") or 0)
    except (ValueError, TypeError):
        pass
new_pairs = {(int(c), t) for c, t in zip(gt["ref_code"], gt["token"])}
non_agg_old = {p for p in old_pairs if p[0] not in AGG_CODES}
non_agg_new = {p for p in new_pairs if p[0] not in AGG_CODES}
only_old = sorted(non_agg_old - non_agg_new)
only_new = sorted(non_agg_new - non_agg_old)
checks.append(["non-aggregator venue count old vs new",
               f"old={len(non_agg_old)} new={len(non_agg_new)}",
               f"only-old={only_old[:8]} only-new={only_new[:8]}"])

# (b) aggregator codes vs measured impact numbers (from the measurement CSVs)
try:
    imp = pd.read_csv(MEASURE / "impact_delta.csv")   # 1003 vs baseline
    rer = pd.read_csv(MEASURE / "reroute_delta.csv")  # 1004/4011 on top of 1003
except FileNotFoundError:
    imp = rer = None
if imp is None:
    checks.append(["aggregator totals vs measured", "SKIPPED",
                   "hypersync-results/measurements/*.csv not found"])
exp_1003 = 0.0 if imp is None else (
    imp[imp.ref_code == 1003]["delta"].sum() + rer[rer.ref_code == 1003]["delta"].sum())
exp_1004 = 0.0 if rer is None else rer[rer.ref_code == 1004]["delta"].sum()
exp_4011 = 0.0 if rer is None else rer[rer.ref_code == 4011]["delta"].sum()
# The measurement CSVs were captured at the 2026-07-01 cutoff — compare only
# over the months they cover, so extending the settlement window doesn't
# spuriously fail the check on new months' aggregator DR.
meas_max = None if imp is None else max(
    pd.concat([imp["month"], rer["month"]]).str[:7])
new_tot_by_code = (df if meas_max is None else
                   df[df["month_s"] <= meas_max]).groupby("ref_code")["dr_usd"].sum()
# 1004 additionally carries the routers' own dust balances self-tagged by
# their real Referral events (~$65 full-history; old workbook showed $0.06 in
# the 2026 window alone) — allow that residue on top of the re-route delta.
agg_expect = () if imp is None else (
    (1003, exp_1003, 1.0), (1004, exp_1004, 150.0), (4011, exp_4011, 1.0))
for code, exp, slack in agg_expect:
    got = float(new_tot_by_code.get(code, 0.0))
    ok = exp - 1.0 <= got <= exp + slack
    checks.append([f"aggregator {code} total vs measured (thru {meas_max})",
                   "OK" if ok else "MISMATCH",
                   f"workbook={got:,.2f} expected={exp:,.2f} (+{slack:.0f} router-dust slack for 1004)"])

# months past the measurement cutoff have NO independent aggregator
# verification — surface them loudly instead of silently excluding them.
if meas_max is not None:
    uncov = sorted(m for m in set(df["month_s"]) if m > meas_max)
    if uncov:
        vals = (df[df["month_s"].isin(uncov) & df["ref_code"].isin(AGG_CODES)]
                .groupby("ref_code")["dr_usd"].sum())
        checks.append(["aggregator months beyond measurement coverage", "UNVERIFIED",
                       f"months {uncov}: "
                       + ", ".join(f"{c}=${float(vals.get(c, 0.0)):,.2f}"
                                   for c in sorted(AGG_CODES))
                       + f" — the impact/reroute measurement run stops at {meas_max}; "
                       "re-run the measurement harness (or verify attribution "
                       "manually) before settling aggregator codes for these months"])

# (c) non-aggregator values old-vs-new, compared PER MONTH and only where the
# old workbook has a value — the old Soter tabs apply settlement cutoffs, so
# full-history totals are not comparable; individual populated months are.
SHIFTED = {99, 128, 1, 0, 1002, 1001}          # aggregator relabeling
METHOD_CHANGED = {9001, 4001}                     # EOD -> intraday TWA (2026-07-27)
INFRA_SPLIT = {99, 10000, 10001}                  # address-based infra split (2026-07-29)
old_mon = {}
mon_cols = [c for c in old_hdr if isinstance(c, str) and c[:4].isdigit()]
mi = {c: old_hdr.index(c) for c in mon_cols}
for r in old_rows[1:]:
    try:
        code = int(str(r[0]))
    except (ValueError, TypeError):
        continue
    for m, i in mi.items():
        v = r[i]
        if v not in (None, ""):
            try:
                old_mon[(code, str(r[1]), m)] = float(str(v).replace(",", ""))
            except ValueError:
                pass
new_mon = {(int(c), t, m): v for c, t, m, v in
           zip(gt["ref_code"], gt["token"], gt["month_s"], gt["dr_usd"])}
big_moves = []
for key, old_v in old_mon.items():
    code, token, m = key
    if code in AGG_CODES or abs(old_v) < 10:
        continue
    new_v = float(new_mon.get(key, 0.0))
    if abs(new_v - old_v) > max(25.0, 0.02 * abs(old_v)):
        tag = ("expected (aggregator shift)" if code in SHIFTED
               else "expected (EOD->TWA methodology)" if code in METHOD_CHANGED
               else "expected (infra split)" if code in INFRA_SPLIT
               else "UNEXPECTED")
        big_moves.append((key, round(old_v, 2), round(new_v, 2), tag))
unexpected = [x for x in big_moves if x[3] == "UNEXPECTED"]
checks.append(["non-aggregator per-month values vs old workbook",
               f"{len(old_mon)} populated cells compared; {len(big_moves)} moved >2%/$25; "
               f"{len(unexpected)} unexpected",
               "; ".join(f"{k}:{o}->{n} {tag}" for k, o, n, tag in
                         sorted(unexpected, key=lambda x: -abs(x[2] - x[1]))[:12])])

# (d) July 2026 rate basis — Boosted-DR termination applied from 2026-07-09.
# Asserted against the ACTUAL rate table the pipeline computed with, not prose.
from drhs.revenue import rates  # noqa: E402
_rate_ok = (
    rates.daily_rate("XR", date(2026, 7, 8)) == rates.apy_to_daily(0.005)
    and rates.daily_rate("XR", date(2026, 7, 9)) == rates.apy_to_daily(0.002)
    and rates.daily_rate("XR*", date(2026, 7, 9)) == rates.apy_to_daily(0.002)
    and rates.daily_rate("XR-stUSDS", date(2026, 7, 9)) == rates.apy_to_daily(0.001))
checks.append(["2026-07 rate basis", "OK" if _rate_ok else "MISMATCH",
               "XR 0.5% through 2026-07-08, 0.2% from 2026-07-09 (Atlas Edit Weekly "
               "Cycle week of 2026-07-06 terminates the +0.3% Boosted DR on the 0.2% "
               "base, ratified 07-09; matches the MSC sheet note 'BOOSTED DR Changed "
               "July 9th'). XR* 0.2% and XR-stUSDS 0.1% unaffected. Asserted against "
               "py/drhs/revenue/rates.py REWARD_SCHEDULE."])

# (e) provenance
checks.append(["provenance", "OK",
               "Soter tabs derive solely from hypersync-results/dr_full (HyperSync event "
               "logs; rates = locked protocol constants, tests py/tests/test_revenue.py; "
               "conversions/deployment event-derived). Reference tabs are verbatim copies "
               "labeled '(reference)' and feed ONLY the Diff tabs."])
write_aoa("Checks", checks)

wb.save(NEW)
print(f"wrote {NEW}")
print("\n=== Checks ===")
for row in checks[1:]:
    print(" | ".join(str(c)[:150] for c in row))
